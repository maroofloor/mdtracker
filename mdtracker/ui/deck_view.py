"""덱 관리 뷰 — 대표 카드 아트 + 체크박스 다중선택 삭제 + 전체삭제 + 엑셀 내보내기/가져오기."""

from __future__ import annotations

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..cardart import archetypes
from ..cardart.service import CardArtService
from ..styles.theme import (
    ACCENT, BORDER, IMPORT_COLOR, LOSE_DARK, PANEL,
    SURFACE, SURFACE2, TEXT, TEXT2, WIN_DARK,
)
from .art_fetch import ArtFetcher
from .completer import make_deck_completer
from .deck_avatar import DeckAvatar

_BTN = "border-radius: 8px; padding: 6px 14px; font-weight: bold; color: #fff; border: none;"


def _centered(widget: QWidget) -> QWidget:
    """셀 내부에서 위젯을 가운데 정렬하는 투명 컨테이너."""
    box = QWidget()
    box.setStyleSheet("background: transparent;")
    lay = QHBoxLayout(box)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.addStretch()
    lay.addWidget(widget)
    lay.addStretch()
    return box


class DeckView(QWidget):
    def __init__(self, db) -> None:
        super().__init__()
        self.db = db
        self.art = CardArtService(db.decks)
        self._avatars: dict[str, DeckAvatar] = {}
        self._gallery_avatars: dict[str, DeckAvatar] = {}
        self._fetcher = ArtFetcher(self.art, self)
        self._fetcher.fetched.connect(self._on_art_fetched)
        self._loading = False
        self._build()
        self.refresh()

    def _build(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # ── 검색 ──────────────────────────────────────────────────
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍  덱 이름 검색...")
        self.search.setStyleSheet(
            f"background: {SURFACE}; color: {TEXT}; border: 1px solid {BORDER};"
            "border-radius: 8px; padding: 6px 12px; font-size: 13px;")
        self.search.textChanged.connect(self._filter)
        root.addWidget(self.search)

        # ── 추가 행 ───────────────────────────────────────────────
        add_row = QHBoxLayout()
        self.name_edit = QLineEdit()
        self.name_edit.setPlaceholderText("새 덱 이름 (예: 스네이크아이)")
        self.name_edit.setStyleSheet(
            f"background: {SURFACE}; color: {TEXT}; border: 1px solid {BORDER};"
            "border-radius: 8px; padding: 6px 12px;")
        self.name_edit.setCompleter(make_deck_completer(
            self.art.db.all_display_names()
            + [d.name for d in self.db.decks.list()], self))
        self.add_btn = QPushButton("+ 추가")
        self.add_btn.setStyleSheet(f"background: {ACCENT}; {_BTN}")
        self.mine_check = QCheckBox("내 덱")
        self.mine_check.setChecked(True)
        self.mine_check.setStyleSheet(f"color: {TEXT2};")
        add_row.addWidget(self.name_edit, 1)
        add_row.addWidget(self.mine_check)
        add_row.addWidget(self.add_btn)
        root.addLayout(add_row)

        # ── 안내 텍스트 ───────────────────────────────────────────
        hint = QLabel(
            "아트 더블클릭 → 대표 카드 지정 · '내 덱' 체크로 내 덱 설정 · "
            "'선택' 체크 후 선택 삭제")
        hint.setStyleSheet(f"color: {TEXT2}; font-size: 11px;")
        root.addWidget(hint)

        # ── 테이블 (아트 / 선택 / 덱 이름 / 내 덱) ────────────────
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["아트", "선택", "덱 이름", "내 덱"])
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)
        header.setSectionResizeMode(1, QHeaderView.Fixed)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Fixed)
        self.table.setColumnWidth(0, 52)
        self.table.setColumnWidth(1, 52)
        self.table.setColumnWidth(3, 72)
        self.table.verticalHeader().hide()
        self.table.setShowGrid(False)
        self.table.setStyleSheet(
            f"QTableWidget {{ background: {PANEL}; border: 1px solid {BORDER}; }}"
            f"QTableWidget::item {{ padding: 6px 12px; color: {TEXT}; }}"
            f"QHeaderView::section {{ background: {SURFACE}; color: {TEXT2};"
            f"font-size: 11px; border: none; border-bottom: 1px solid {BORDER};"
            "padding: 6px 12px; }}")
        self.table.verticalHeader().setDefaultSectionSize(46)
        root.addWidget(self.table, 1)

        # ── 갤러리 (아트 카드 그리드) ─────────────────────────────
        self._gallery = QScrollArea()
        self._gallery.setWidgetResizable(True)
        self._gallery.setFrameShape(QScrollArea.Shape.NoFrame)
        self._gallery.setStyleSheet("background: transparent;")
        gal_container = QWidget()
        gal_container.setStyleSheet("background: transparent;")
        self._gallery_grid = QGridLayout(gal_container)
        self._gallery_grid.setContentsMargins(4, 4, 4, 4)
        self._gallery_grid.setSpacing(12)
        self._gallery_grid.setAlignment(
            Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self._gallery.setWidget(gal_container)
        self._gallery.hide()
        root.addWidget(self._gallery, 1)

        from .empty_state import EmptyState
        self.empty_state = EmptyState(
            icon="🗂", title="등록된 덱이 없습니다",
            subtitle="위에 덱 이름을 입력해 추가하거나 ‘🎴 자동 매핑’으로 "
                     "아트를 가져오세요.")
        self.empty_state.hide()
        root.addWidget(self.empty_state, 1)

        # ── 하단 버튼 ─────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self.gallery_btn = QPushButton("🖼 갤러리")
        self.gallery_btn.setCheckable(True)
        self.gallery_btn.setStyleSheet(f"background: {SURFACE2}; {_BTN}")
        btn_row.addWidget(self.gallery_btn)
        self.auto_btn = QPushButton("🎴 자동 매핑")
        self.auto_btn.setStyleSheet(f"background: {ACCENT}; {_BTN}")
        self.auto_btn.setToolTip(
            "매핑표에 있는 덱의 대표 카드 아트를 자동으로 가져옵니다(네트워크).")
        self.del_btn = QPushButton("선택 삭제")
        self.del_btn.setStyleSheet(f"background: #ef4444; {_BTN}")
        self.del_all_btn = QPushButton("전체 삭제")
        self.del_all_btn.setStyleSheet(f"background: {LOSE_DARK}; {_BTN}")
        btn_row.addWidget(self.auto_btn)
        btn_row.addWidget(self.del_btn)
        btn_row.addWidget(self.del_all_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)

        self.gallery_btn.toggled.connect(self._on_toggle_gallery)
        self.auto_btn.clicked.connect(self._on_auto_map)
        self.add_btn.clicked.connect(self._on_add)
        self.name_edit.returnPressed.connect(self._on_add)
        self.del_btn.clicked.connect(self._on_delete)
        self.del_all_btn.clicked.connect(self._on_delete_all)
        self.table.itemChanged.connect(self._on_item_changed)

    def refresh(self) -> None:
        self._loading = True
        try:
            self._avatars = {}
            decks = self.db.decks.list()   # is_mine DESC, name ASC
            self.table.setRowCount(len(decks))
            for row, d in enumerate(decks):
                row_bg = QColor(SURFACE) if row % 2 else QColor(SURFACE2)

                # 아트 아바타 (더블클릭 → 대표 카드 지정)
                avatar = DeckAvatar(d.name, self.art.local_path(d.name), size=34)
                avatar.clicked.connect(
                    lambda n=d.name: self._assign_art(n))
                self.table.setCellWidget(row, 0, _centered(avatar))
                self._avatars[d.name] = avatar
                # 아트가 없고 매핑되는 덱이면 백그라운드로 자동 로드
                self._fetcher.request(d.name)

                # 선택 체크박스 (편집 불가)
                sel_item = QTableWidgetItem()
                sel_item.setCheckState(Qt.Unchecked)
                sel_item.setTextAlignment(Qt.AlignCenter)
                sel_item.setBackground(row_bg)
                sel_item.setFlags(sel_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 1, sel_item)

                # 덱 이름 (더블클릭 인라인 편집 가능)
                name_item = QTableWidgetItem(d.name)
                name_item.setData(Qt.UserRole, d.id)
                name_item.setForeground(QColor(TEXT))
                name_item.setBackground(row_bg)
                name_item.setToolTip("더블클릭하여 이름 편집")
                self.table.setItem(row, 2, name_item)

                # 내 덱 체크박스 (편집 불가 — 클릭으로 토글)
                mine_item = QTableWidgetItem()
                mine_item.setCheckState(Qt.Checked if d.is_mine else Qt.Unchecked)
                mine_item.setTextAlignment(Qt.AlignCenter)
                mine_item.setBackground(row_bg)
                mine_item.setFlags(
                    mine_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, 3, mine_item)
        finally:
            self._loading = False
        self._filter(self.search.text())
        empty = self.table.rowCount() == 0
        self.empty_state.setVisible(empty)
        gallery_on = self.gallery_btn.isChecked()
        self.table.setVisible(not empty and not gallery_on)
        self._gallery.setVisible(not empty and gallery_on)

    def _filter(self, text: str) -> None:
        q = text.strip().lower()
        for row in range(self.table.rowCount()):
            item = self.table.item(row, 2)
            hide = bool(q and item and q not in item.text().lower())
            self.table.setRowHidden(row, hide)
        if getattr(self, "_gallery", None) is not None and \
                self._gallery.isVisible():
            self._build_gallery()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._loading:
            return
        col = item.column()
        if col == 3:
            # 내 덱 체크박스 토글
            id_item = self.table.item(item.row(), 2)
            if not id_item:
                return
            deck_id = id_item.data(Qt.UserRole)
            if deck_id is None:
                return
            self.db.decks.set_mine(deck_id, item.checkState() == Qt.Checked)
            self.refresh()
        elif col == 2:
            # 덱 이름 인라인 편집
            deck_id = item.data(Qt.UserRole)
            new_name = item.text().strip()
            if deck_id is None or not new_name:
                self.refresh()
                return
            self.db.decks.rename(deck_id, new_name)
            self.refresh()

    def _checked_ids(self) -> list[int]:
        ids = []
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            sel_item = self.table.item(row, 1)
            if sel_item and sel_item.checkState() == Qt.Checked:
                name_item = self.table.item(row, 2)
                if name_item:
                    did = name_item.data(Qt.UserRole)
                    if did is not None:
                        ids.append(did)
        return ids

    # ── 카드 아트 ────────────────────────────────────────────────────

    def _on_art_fetched(self, deck_name: str, path: str) -> None:
        if not path:
            return
        for amap in (self._avatars, self._gallery_avatars):
            av = amap.get(deck_name)
            if av is not None:
                av.set_image(path)

    # ── 갤러리 ───────────────────────────────────────────────────────

    def _on_toggle_gallery(self, on: bool) -> None:
        self.gallery_btn.setText("📋 목록" if on else "🖼 갤러리")
        self.table.setVisible(not on)
        self._gallery.setVisible(on)
        self.del_btn.setEnabled(not on)   # 선택 삭제는 목록 모드 전용
        if on:
            self._build_gallery()

    def _build_gallery(self) -> None:
        while self._gallery_grid.count():
            item = self._gallery_grid.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()
        self._gallery_avatars = {}
        q = self.search.text().strip().lower()
        decks = [d for d in self.db.decks.list()
                 if not q or q in d.name.lower()]
        vw = self._gallery.viewport().width()
        cols = max(1, vw // 132)
        for i, d in enumerate(decks):
            self._gallery_grid.addWidget(
                self._make_gallery_card(d), i // cols, i % cols)

    def _make_gallery_card(self, d) -> QFrame:
        card = QFrame()
        card.setFixedSize(120, 156)
        card.setStyleSheet(
            f"QFrame {{ background: {PANEL}; border: 1px solid {BORDER};"
            "border-radius: 10px; }}")
        v = QVBoxLayout(card)
        v.setContentsMargins(8, 10, 8, 8)
        v.setSpacing(6)

        avatar = DeckAvatar(d.name, self.art.local_path(d.name), size=84)
        avatar.clicked.connect(lambda n=d.name: self._assign_art(n))
        self._gallery_avatars[d.name] = avatar
        self._fetcher.request(d.name)
        arow = QHBoxLayout()
        arow.setContentsMargins(0, 0, 0, 0)
        arow.addStretch()
        arow.addWidget(avatar)
        arow.addStretch()
        v.addLayout(arow)

        name = QLabel(d.name)
        name.setAlignment(Qt.AlignmentFlag.AlignCenter)
        name.setWordWrap(True)
        name.setStyleSheet(
            f"color:{TEXT}; font-size:12px; font-weight:600; border:none;"
            "background:transparent;")
        v.addWidget(name)

        if d.is_mine:
            mine = QLabel("내 덱")
            mine.setAlignment(Qt.AlignmentFlag.AlignCenter)
            mine.setStyleSheet(
                f"color:{ACCENT}; font-size:10px; font-weight:700; border:none;"
                "background:transparent;")
            v.addWidget(mine)
        v.addStretch()
        return card

    def _assign_art(self, deck_name: str) -> None:
        dlg = _CardSearchDialog(self, self.art, deck_name)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        card = dlg.selected_card()
        if not card:
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            path = self.art.assign_manual(deck_name, card)
        finally:
            QApplication.restoreOverrideCursor()
        if not path:
            QMessageBox.warning(
                self, "지정 실패",
                "카드 이미지를 내려받지 못했습니다. 네트워크를 확인해주세요.")
        self.refresh()

    def _on_auto_map(self) -> None:
        decks = self.db.decks.list()
        targets = [d for d in decks
                   if self.art.local_path(d.name) is None
                   and self.art.has_auto_mapping(d.name)]
        if not targets:
            QMessageBox.information(
                self, "자동 매핑",
                "자동 매핑할 덱이 없습니다.\n"
                "(이미 아트가 있거나, 매핑표에 없는 덱입니다. "
                "아트를 더블클릭해 수동 지정할 수 있습니다.)")
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        done = 0
        try:
            for d in targets:
                if self.art.ensure(d.name):
                    done += 1
        finally:
            QApplication.restoreOverrideCursor()
        self.refresh()
        QMessageBox.information(
            self, "자동 매핑 완료",
            f"{done}개 덱에 카드 아트를 적용했습니다. (대상 {len(targets)}개)")

    # ── 덱 CRUD ──────────────────────────────────────────────────────

    def _on_add(self) -> None:
        name = self.name_edit.text().strip()
        if not name:
            return
        self.db.decks.add(name, is_mine=self.mine_check.isChecked())
        self.name_edit.clear()
        self.refresh()

    def _on_delete(self) -> None:
        ids = self._checked_ids()
        if not ids:
            QMessageBox.information(self, "선택 없음", "삭제할 덱을 '선택' 열에서 체크해주세요.")
            return
        yes = QMessageBox.StandardButton.Yes
        if QMessageBox.question(
            self, "선택 삭제",
            f"체크된 덱 {len(ids)}개를 삭제할까요?\n(기존 대전 기록은 영향받지 않습니다)",
            yes | QMessageBox.StandardButton.No,
        ) == yes:
            for did in ids:
                self.db.decks.delete(did)
            self.refresh()

    def _on_delete_all(self) -> None:
        count = self.table.rowCount()
        if count == 0:
            return
        yes = QMessageBox.StandardButton.Yes
        if QMessageBox.question(
            self, "전체 삭제",
            f"덱 목록 전체 {count}개를 모두 삭제할까요?\n"
            "이 작업은 되돌릴 수 없습니다. 기존 대전 기록은 영향받지 않습니다.",
            yes | QMessageBox.StandardButton.No,
        ) == yes:
            self.db.decks.delete_all()
            self.refresh()

    def _on_export(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "덱 목록 저장", "decks.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "덱 목록"
            ws.append(["덱 이름", "내 덱"])
            decks = self.db.decks.list()
            for d in decks:
                ws.append([d.name, "O" if d.is_mine else ""])
            wb.save(path)
            QMessageBox.information(self, "완료", f"{len(decks)}개 덱을 저장했습니다.")
        except ImportError:
            QMessageBox.warning(
                self, "오류",
                "openpyxl이 설치되지 않았습니다.\n.venv/bin/pip install openpyxl")
        except Exception as e:
            QMessageBox.warning(self, "저장 실패", str(e))

    def _on_import(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "엑셀 파일 열기", "", "Excel / CSV (*.xlsx *.xls *.csv)")
        if not path:
            return
        try:
            rows = _read_deck_file(path)
        except Exception as e:
            QMessageBox.warning(self, "가져오기 실패", str(e))
            return
        if not rows:
            QMessageBox.information(self, "가져오기", "가져올 덱이 없습니다.")
            return
        added = sum(1 for name, is_mine in rows
                    if name and self.db.decks.add(name, is_mine=is_mine))
        self.refresh()
        QMessageBox.information(self, "완료", f"{added}개 덱을 가져왔습니다.")


class _CardSearchDialog(QDialog):
    """YGOPRODeck 카드명 검색 → 대표 카드 선택 다이얼로그(수동 보정)."""

    def __init__(self, parent, art_service: CardArtService,
                 deck_name: str) -> None:
        super().__init__(parent)
        self._art = art_service
        self.setWindowTitle(f"대표 카드 지정 — {deck_name}")
        self.resize(420, 460)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        info = QLabel("카드명(영문/일부)으로 검색해 대표 카드를 고르세요.")
        info.setStyleSheet(f"color: {TEXT2}; font-size: 12px;")
        lay.addWidget(info)

        search_row = QHBoxLayout()
        self.edit = QLineEdit()
        # 매핑표에 영문 아키타입이 있으면 프리필
        self.edit.setText(archetypes.resolve_archetype(deck_name) or deck_name)
        self.edit.setStyleSheet(
            f"background: {SURFACE}; color: {TEXT}; border: 1px solid {BORDER};"
            "border-radius: 6px; padding: 6px 10px;")
        self.edit.returnPressed.connect(self._do_search)
        search_btn = QPushButton("검색")
        search_btn.setStyleSheet(f"background: {ACCENT}; {_BTN}")
        search_btn.clicked.connect(self._do_search)
        search_row.addWidget(self.edit, 1)
        search_row.addWidget(search_btn)
        lay.addLayout(search_row)

        self.results = QListWidget()
        self.results.setStyleSheet(
            f"background: {PANEL}; color: {TEXT}; border: 1px solid {BORDER};"
            "border-radius: 6px;")
        self.results.itemDoubleClicked.connect(lambda _i: self.accept())
        lay.addWidget(self.results, 1)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok
            | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        lay.addWidget(buttons)

    def _do_search(self) -> None:
        term = self.edit.text().strip()
        if not term:
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            cards = self._art.search(term)
        finally:
            QApplication.restoreOverrideCursor()
        self.results.clear()
        if not cards:
            QMessageBox.information(
                self, "검색 결과 없음",
                "결과가 없습니다. 영문 카드명으로 다시 시도해보세요.")
            return
        for c in cards:
            item = QListWidgetItem(c.get("name", "(이름 없음)"))
            item.setData(Qt.UserRole, c)
            self.results.addItem(item)
        self.results.setCurrentRow(0)

    def selected_card(self) -> dict | None:
        item = self.results.currentItem()
        return item.data(Qt.UserRole) if item else None


def _read_deck_file(path: str) -> list[tuple[str, bool]]:
    """xlsx / xls / csv 에서 (덱 이름, is_mine) 목록을 읽는다."""
    if path.lower().endswith(".csv"):
        import csv
        with open(path, encoding="utf-8-sig", newline="") as f:
            raw = list(csv.reader(f))
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw = [[str(c.value or "") for c in r] for r in ws.iter_rows()]

    result: list[tuple[str, bool]] = []
    for i, row in enumerate(raw):
        name = row[0].strip() if row else ""
        if i == 0 and name.lower() in ("덱 이름", "덱이름", "deck", "name"):
            continue  # 헤더 행 스킵
        if not name:
            continue
        mine_raw = row[1].strip().upper() if len(row) > 1 else ""
        is_mine = mine_raw in ("O", "TRUE", "1", "Y", "YES", "내덱")
        result.append((name, is_mine))
    return result
